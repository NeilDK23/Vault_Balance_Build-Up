from __future__ import annotations

# These imports are built-in Python tools used for reading CSV files,
# cleaning text, handling dates, and working with folder/file paths.
import csv
import re
from datetime import datetime
from pathlib import Path
from typing import Iterable

# openpyxl is the Excel library used to create, format, and save the workbook.
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


# -----------------------------------------------------------------------------
# FOLDER PATHS
# These tell the script where the project lives and where to read and write files.
# -----------------------------------------------------------------------------
BASE_DIR = Path(r"G:\Shared drives\Lumepay (Restricted)\Accounting\Miscellaneous\Accounting Devs\Fireblocks Vault Build-Up")
CSV_DIR = BASE_DIR / "Vault CSV"
OUTPUT_DIR = BASE_DIR / "Output"
RECON_DIR = BASE_DIR / "Recon Vault Report"


# -----------------------------------------------------------------------------
# ASSET GROUPS
# These lists/sets decide which currencies are included in each worksheet.
# -----------------------------------------------------------------------------
ALLOWED_ASSETS = ["TRX", "ETH", "USDT_ERC20", "TRX_USDT_S2UZ"]
USDT_ASSETS = {"USDT_ERC20", "TRX_USDT_S2UZ"}
TRX_SHEET_ASSETS = {"TRX", "TRX_USDT_S2UZ"}
ETH_SHEET_ASSETS = {"ETH", "USDT_ERC20"}
RECON_ASSETS = ["ETH", "TRX", "USDT_ERC20", "TRX_USDT_S2UZ"]


# -----------------------------------------------------------------------------
# EXCEL FORMATS
# These control how dates, numbers, percentages, widths, and fill colours appear.
# -----------------------------------------------------------------------------
DATE_NUMBER_FORMAT = "yyyy/mm/dd hh:mm"
ROUNDED_DATE_NUMBER_FORMAT = "yyyy/mm/dd"
COMMA_STYLE_FORMAT = "#,##0.00"
TABLE_COLUMN_WIDTH = 16.5
PERCENT_FORMAT = "0.0000%"

BLUE_HEADER_FILL = PatternFill(fill_type="solid", fgColor="538DD5")
BLUE_DATA_FILL = PatternFill(fill_type="solid", fgColor="C5D9F1")
YELLOW_HEADER_FILL = PatternFill(fill_type="solid", fgColor="FFE599")
YELLOW_DATA_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")


# -----------------------------------------------------------------------------
# FILE FINDING HELPERS
# These functions locate the input files the script expects to use.
# -----------------------------------------------------------------------------
def list_csv_files(folder: Path) -> list[Path]:
    # Find all CSV files in the Vault CSV folder.
    csv_files = sorted(folder.glob("*.csv"))

    # The script is built around exactly 2 Fireblocks exports: source and destination.
    if len(csv_files) != 2:
        raise ValueError(f"Expected exactly 2 CSV files in '{folder}', found {len(csv_files)}.")

    return csv_files



def find_destination_csv(csv_files: list[Path]) -> Path:
    # The destination export is identified by having "Destination" in the filename.
    for csv_file in csv_files:
        if "destination" in csv_file.name.lower():
            return csv_file

    raise ValueError("Could not find the CSV file with 'Destination' in the filename.")



def find_recon_file(folder: Path) -> Path:
    # The Recon folder should contain only one report file.
    files = [path for path in folder.iterdir() if path.is_file()]
    if len(files) != 1:
        raise ValueError(f"Expected exactly 1 recon file in '{folder}', found {len(files)}.")
    return files[0]


# -----------------------------------------------------------------------------
# FILE READING HELPERS
# These functions load CSV data into Python lists.
# -----------------------------------------------------------------------------
def read_destination_rows(destination_csv: Path) -> list[list[str]]:
    with destination_csv.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        rows = list(reader)

    # We expect a header row plus at least one row of data.
    if len(rows) < 2:
        raise ValueError(f"Destination CSV does not contain a data row: {destination_csv}")

    return rows



def read_recon_rows(recon_file: Path) -> list[list[str]]:
    with recon_file.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        rows = list(reader)

    if not rows:
        raise ValueError(f"Recon file is empty: {recon_file}")

    return rows



def read_trimmed_rows(csv_path: Path) -> tuple[list[str], list[list[str]]]:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        rows = list(reader)

    if not rows:
        raise ValueError(f"CSV file is empty: {csv_path}")

    # Keep only the first 29 columns, which corresponds to Excel columns A:AC.
    header = rows[0][:29]

    # Ignore completely blank rows when importing.
    data_rows = [row[:29] for row in rows[1:] if any(cell.strip() for cell in row)]
    return header, data_rows


# -----------------------------------------------------------------------------
# VAULT DETAILS
# The destination CSV also tells us the vault name and the wallet addresses we
# want to display on the sheets.
# -----------------------------------------------------------------------------
def read_vault_details(destination_csv: Path) -> tuple[str, str, str]:
    rows = read_destination_rows(destination_csv)
    first_data_row = rows[1]

    # Column R in the destination export contains the vault name.
    vault_name = first_data_row[17].strip()
    if not vault_name:
        raise ValueError("Vault name in column R of the Destination CSV is blank.")

    trx_address = ""
    eth_address = ""

    # Look through the destination file until we find the TRX and ETH wallet addresses.
    for row in rows[1:]:
        if len(row) <= 18:
            continue

        asset = row[4].strip()
        destination_address = row[18].strip()

        if asset == "TRX" and not trx_address:
            trx_address = destination_address
        if asset == "ETH" and not eth_address:
            eth_address = destination_address

        # Stop early once both addresses have been found.
        if trx_address and eth_address:
            break

    if not trx_address:
        raise ValueError("Could not find a TRX destination address in the Destination CSV.")
    if not eth_address:
        raise ValueError("Could not find an ETH destination address in the Destination CSV.")

    return vault_name, trx_address, eth_address



def build_output_file(vault_name: str) -> Path:
    # Remove characters that are not safe in a Windows filename.
    safe_name = re.sub(r'[<>:"/\\|?*]', "_", vault_name).strip() or "Vault"
    return OUTPUT_DIR / f"{safe_name} Asset Build-Up.xlsx"


# -----------------------------------------------------------------------------
# DATA CLEANING AND FILTERING
# These functions convert raw text into useful values and keep only the rows we want.
# -----------------------------------------------------------------------------
def parse_fireblocks_date(raw_value: str) -> datetime:
    # Fireblocks dates look like "05 Apr 2026 08:32:53 GMT".
    # We only need the date/time part, not the timezone text.
    trimmed = raw_value[:20].strip()
    return datetime.strptime(trimmed, "%d %b %Y %H:%M:%S")



def filter_completed(rows: Iterable[list[str]]) -> list[list[str]]:
    # Keep only rows where Status (column B) is COMPLETED.
    return [row for row in rows if len(row) > 1 and row[1].strip().upper() == "COMPLETED"]



def filter_assets(rows: Iterable[list]) -> list[list]:
    # Keep only the assets we care about for this workbook.
    return [row for row in rows if len(row) > 4 and str(row[4]).strip() in ALLOWED_ASSETS]



def to_decimal_amount(value) -> float:
    # Convert text like "123.45" into a real number.
    # If the cell is blank, treat it as 0.
    if value in (None, ""):
        return 0.0
    return float(value)



def convert_numeric_columns(row: list[str]) -> list:
    # Create a copy of the row so we do not accidentally change the original list.
    converted_row = list(row)

    # Columns H to M in Excel are numeric columns, so convert them from text to numbers.
    for index in range(7, 13):
        converted_row[index] = to_decimal_amount(converted_row[index])

    return converted_row


# -----------------------------------------------------------------------------
# GENERAL EXCEL FORMATTING
# These functions apply widths, filters, freezing, number formats, and fills.
# -----------------------------------------------------------------------------
def set_table_column_widths(worksheet, max_column: int) -> None:
    # Set all columns in the table area to the same width for consistency.
    for column_idx in range(1, max_column + 1):
        worksheet.column_dimensions[get_column_letter(column_idx)].width = TABLE_COLUMN_WIDTH



def apply_table_formats(worksheet, header_row: int, first_data_row: int, max_column: int) -> None:
    # Make the header row bold.
    for cell in worksheet[header_row]:
        cell.font = Font(bold=True)

    # Turn on Excel filters for the whole table range.
    worksheet.auto_filter.ref = f"A{header_row}:{get_column_letter(max_column)}{worksheet.max_row}"

    # Freeze the screen just below the header so headings stay visible while scrolling.
    worksheet.freeze_panes = f"A{first_data_row}"

    # Apply comma-style number formatting to columns H:M.
    for row_idx in range(first_data_row, worksheet.max_row + 1):
        for column_idx in range(8, 14):
            worksheet.cell(row=row_idx, column=column_idx).number_format = COMMA_STYLE_FORMAT

    # Apply date formatting to the added Date and Date rounded columns.
    for row_idx in range(first_data_row, worksheet.max_row + 1):
        worksheet.cell(row=row_idx, column=30).number_format = DATE_NUMBER_FORMAT
        worksheet.cell(row=row_idx, column=31).number_format = ROUNDED_DATE_NUMBER_FORMAT

    set_table_column_widths(worksheet, max_column)



def style_top_label_rows(worksheet) -> None:
    # On the asset sheets, rows 1 and 2 contain labels that should stand out.
    worksheet["A1"].font = Font(bold=True)
    worksheet["A2"].font = Font(bold=True)



def apply_colour_block(worksheet, header_row: int, first_data_row: int, start_col: int, end_col: int, header_fill, data_fill) -> None:
    # Apply one colour to the header row and a different colour to the data below it.
    for column_idx in range(start_col, end_col + 1):
        worksheet.cell(row=header_row, column=column_idx).fill = header_fill
        for row_idx in range(first_data_row, worksheet.max_row + 1):
            worksheet.cell(row=row_idx, column=column_idx).fill = data_fill



def apply_usdt_colour_blocks(worksheet, header_row: int, first_data_row: int) -> None:
    # Blue for AD:AE and yellow for AF:AH on the USDT sheet.
    apply_colour_block(worksheet, header_row, first_data_row, 30, 31, BLUE_HEADER_FILL, BLUE_DATA_FILL)
    apply_colour_block(worksheet, header_row, first_data_row, 32, 34, YELLOW_HEADER_FILL, YELLOW_DATA_FILL)



def apply_base_asset_colour_blocks(worksheet, header_row: int, first_data_row: int) -> None:
    # Blue for AD:AE and yellow for AF:AI on the TRX and ETH sheets.
    apply_colour_block(worksheet, header_row, first_data_row, 30, 31, BLUE_HEADER_FILL, BLUE_DATA_FILL)
    apply_colour_block(worksheet, header_row, first_data_row, 32, 35, YELLOW_HEADER_FILL, YELLOW_DATA_FILL)


# -----------------------------------------------------------------------------
# SHEET WRITERS
# These functions place rows into worksheets and apply the common sheet layout.
# -----------------------------------------------------------------------------
def write_standard_sheet(workbook: Workbook, title: str, header: list[str], rows: list[list]) -> None:
    worksheet = workbook.create_sheet(title=title)
    worksheet.append(header)

    for row in rows:
        worksheet.append(row)

    apply_table_formats(worksheet, header_row=1, first_data_row=2, max_column=len(header))



def write_single_wallet_sheet(
    workbook: Workbook,
    title: str,
    header: list[str],
    rows: list[list],
    wallet_label: str,
    wallet_address: str,
    crypto_asset: str,
) -> None:
    worksheet = workbook.create_sheet(title=title)

    # Row 1 shows the wallet label and wallet address.
    worksheet["A1"] = wallet_label
    worksheet["B1"] = wallet_address

    # Row 2 shows the main asset this sheet is focused on.
    worksheet["A2"] = "Crypto Asset"
    worksheet["B2"] = crypto_asset

    # Row 3 is left blank on purpose.
    worksheet.append([])

    # Row 4 contains the table headings.
    worksheet.append(header)

    # Data starts on row 5.
    for row in rows:
        worksheet.append(row)

    style_top_label_rows(worksheet)
    apply_table_formats(worksheet, header_row=4, first_data_row=5, max_column=len(header))



def write_usdt_sheet(
    workbook: Workbook,
    title: str,
    header: list[str],
    rows: list[list],
    vault_name: str,
    trx_address: str,
    eth_address: str,
) -> None:
    worksheet = workbook.create_sheet(title=title)

    # The USDT sheet needs both the TRX and ETH wallet addresses.
    worksheet["A1"] = f"{vault_name} TRX Address"
    worksheet["B1"] = trx_address
    worksheet["A2"] = f"{vault_name} ETH Address"
    worksheet["B2"] = eth_address

    worksheet.append([])
    worksheet.append(header)

    for row in rows:
        worksheet.append(row)

    style_top_label_rows(worksheet)
    apply_table_formats(worksheet, header_row=4, first_data_row=5, max_column=len(header))


# -----------------------------------------------------------------------------
# FORMULA BUILDERS
# These functions build rows with Excel formulas for balances and fees.
# -----------------------------------------------------------------------------
def build_usdt_rows(rows: list[list]) -> tuple[list[str], list[list]]:
    # Add the three calculated columns for the USDT sheet.
    header = rows[0][:31] + ["Opening balance", "Inflow/(Outflow)", "Closing balance"]

    # Only keep the assets that belong on the USDT sheet.
    usdt_source_rows = [row[:31] for row in rows[1:] if str(row[4]).strip() in USDT_ASSETS]
    output_rows: list[list] = []

    # Data begins on row 5 in Excel, so formulas must also start at row 5.
    for excel_row_number, row in enumerate(usdt_source_rows, start=5):
        # First opening balance is 0, then each later row carries forward the prior closing balance.
        opening_balance = 0 if excel_row_number == 5 else f"=AH{excel_row_number - 1}"

        # If the destination address matches one of the known vault addresses, it is an inflow.
        inflow_outflow = f"=IF(OR(S{excel_row_number}=$B$1,S{excel_row_number}=$B$2),J{excel_row_number},-J{excel_row_number})"

        # Closing balance is opening balance plus movement.
        closing_balance = f"=AF{excel_row_number}+AG{excel_row_number}"

        output_rows.append(row + [opening_balance, inflow_outflow, closing_balance])

    return header, output_rows



def build_base_asset_rows(rows: list[list], sheet_assets: set[str]) -> tuple[list[str], list[list]]:
    # Add the four calculated columns used by the TRX and ETH sheets.
    header = rows[0][:31] + ["Opening Balance", "Inflow/(Outflow)", "Gas Fees", "Closing Balance"]

    # Only keep the assets that belong on this sheet.
    source_rows = [row[:31] for row in rows[1:] if str(row[4]).strip() in sheet_assets]
    output_rows: list[list] = []

    for excel_row_number, row in enumerate(source_rows, start=5):
        opening_balance = 0 if excel_row_number == 5 else f"=AI{excel_row_number - 1}"

        # Only the base asset (TRX on the TRX sheet, ETH on the ETH sheet) affects inflow/outflow.
        inflow_outflow = f"=IF(F{excel_row_number}=$B$2,IF(S{excel_row_number}=$B$1,J{excel_row_number},-J{excel_row_number}),0)"

        # Gas fees only apply where the source address matches the wallet address shown at the top.
        gas_fees = f"=IF(P{excel_row_number}=$B$1,-L{excel_row_number},0)"

        # Closing balance is opening balance + inflow/outflow + gas fees.
        closing_balance = f"=AF{excel_row_number}+AG{excel_row_number}+AH{excel_row_number}"

        output_rows.append(row + [opening_balance, inflow_outflow, gas_fees, closing_balance])

    return header, output_rows


# -----------------------------------------------------------------------------
# EXTRA NUMBER FORMATTING
# These functions format the formula-driven balance columns.
# -----------------------------------------------------------------------------
def apply_usdt_number_formats(worksheet, first_data_row: int) -> None:
    for row_idx in range(first_data_row, worksheet.max_row + 1):
        for column_idx in range(32, 35):
            worksheet.cell(row=row_idx, column=column_idx).number_format = COMMA_STYLE_FORMAT



def apply_base_asset_number_formats(worksheet, first_data_row: int) -> None:
    for row_idx in range(first_data_row, worksheet.max_row + 1):
        for column_idx in range(32, 36):
            worksheet.cell(row=row_idx, column=column_idx).number_format = COMMA_STYLE_FORMAT


# -----------------------------------------------------------------------------
# RECON SHEET PREPARATION
# These functions prepare and write the reconciliation sheet.
# -----------------------------------------------------------------------------
def build_recon_rows(recon_rows: list[list[str]], vault_name: str) -> list[list]:
    filtered = []

    # Keep only rows for the current vault and the assets we care about.
    for row in recon_rows[1:]:
        if len(row) <= 3:
            continue
        if row[0].strip() != vault_name:
            continue
        if row[2].strip() not in RECON_ASSETS:
            continue

        # In the vault report file, column E contains the total balance.
        filtered.append([row[0].strip(), row[2].strip(), to_decimal_amount(row[4])])

    # Keep the recon rows in a predictable order.
    filtered.sort(key=lambda item: RECON_ASSETS.index(item[1]))

    # Create a combined USDT balance by adding the ERC20 and TRX USDT balances together.
    asset_to_balance = {asset: balance for _, asset, balance in filtered}
    usdt_total = asset_to_balance.get("USDT_ERC20", 0.0) + asset_to_balance.get("TRX_USDT_S2UZ", 0.0)

    standard_rows = [row for row in filtered if row[1] in {"ETH", "TRX"}]
    usdt_component_rows = [row for row in filtered if row[1] in {"USDT_ERC20", "TRX_USDT_S2UZ"}]

    # Final row order: ETH, TRX, USDT, then the two component rows below it.
    ordered_rows = standard_rows + [[vault_name, "USDT", usdt_total]] + usdt_component_rows
    return ordered_rows



def write_recon_sheet(workbook: Workbook, rows: list[list]) -> None:
    worksheet = workbook.create_sheet(title="Recon")

    header = ["Vault / Account Name", "Asset", "Vault Report Balance", "Asset Build-Up", "Difference", "Difference %"]
    worksheet.append(header)

    for row in rows:
        worksheet.append(row)

    # Only TRX, ETH, and USDT participate in the final comparison formulas.
    for row_idx in range(2, worksheet.max_row + 1):
        asset = str(worksheet.cell(row=row_idx, column=2).value)

        if asset == "TRX":
            worksheet.cell(row=row_idx, column=4).value = '=LOOKUP(9.99999999999999E+307,TRX!AI:AI)'
            worksheet.cell(row=row_idx, column=5).value = f"=D{row_idx}-C{row_idx}"
            worksheet.cell(row=row_idx, column=6).value = f'=IF(C{row_idx}=0,"na",E{row_idx}/C{row_idx})'
        elif asset == "ETH":
            worksheet.cell(row=row_idx, column=4).value = '=LOOKUP(9.99999999999999E+307,ETH!AI:AI)'
            worksheet.cell(row=row_idx, column=5).value = f"=D{row_idx}-C{row_idx}"
            worksheet.cell(row=row_idx, column=6).value = f'=IF(C{row_idx}=0,"na",E{row_idx}/C{row_idx})'
        elif asset == "USDT":
            worksheet.cell(row=row_idx, column=4).value = '=LOOKUP(9.99999999999999E+307,USDT!AH:AH)'
            worksheet.cell(row=row_idx, column=5).value = f"=D{row_idx}-C{row_idx}"
            worksheet.cell(row=row_idx, column=6).value = f'=IF(C{row_idx}=0,"na",E{row_idx}/C{row_idx})'

    # Make the header row bold and add the usual table setup.
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    worksheet.auto_filter.ref = f"A1:F{worksheet.max_row}"
    worksheet.freeze_panes = "A2"
    set_table_column_widths(worksheet, 6)

    # Format balances and percentages.
    for row_idx in range(2, worksheet.max_row + 1):
        for column_idx in range(3, 6):
            worksheet.cell(row=row_idx, column=column_idx).number_format = COMMA_STYLE_FORMAT
        worksheet.cell(row=row_idx, column=6).number_format = PERCENT_FORMAT


# -----------------------------------------------------------------------------
# MAIN PROGRAM FLOW
# This is the part that runs the full job from start to finish.
# -----------------------------------------------------------------------------
def main() -> None:
    # Find all input files needed for the workbook.
    csv_files = list_csv_files(CSV_DIR)
    destination_csv = find_destination_csv(csv_files)
    vault_name, trx_address, eth_address = read_vault_details(destination_csv)
    output_file = build_output_file(vault_name)
    recon_file = find_recon_file(RECON_DIR)
    recon_source_rows = read_recon_rows(recon_file)

    # Import the two Fireblocks CSV files and combine them into one list.
    combined_rows: list[list[str]] = []
    trimmed_header: list[str] | None = None

    for csv_file in csv_files:
        header, rows = read_trimmed_rows(csv_file)
        if trimmed_header is None:
            trimmed_header = header
        elif header != trimmed_header:
            raise ValueError(f"CSV headers do not match in file: {csv_file}")
        combined_rows.extend(rows)

    if trimmed_header is None:
        raise ValueError("No CSV header was loaded.")

    # Keep only completed rows.
    completed_rows = filter_completed(combined_rows)

    # Add two new columns: Date and Date rounded.
    extended_header = trimmed_header + ["Date", "Date rounded"]
    enriched_rows: list[list] = []

    for row in completed_rows:
        # Convert number-like columns from text into real numbers.
        converted_row = convert_numeric_columns(row)

        # Build the extra date columns from the original Fireblocks date text.
        parsed_date = parse_fireblocks_date(converted_row[3])
        rounded_date = parsed_date.replace(hour=0, minute=0, second=0, microsecond=0)

        enriched_rows.append(converted_row + [parsed_date, rounded_date])

    # Keep only the assets we want and sort them by the new Date column.
    filtered_rows = filter_assets(enriched_rows)
    filtered_rows.sort(key=lambda row: row[29])

    # Start a new Excel workbook and remove the empty default sheet.
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    # Create the sheets in the required tab order.
    recon_rows = build_recon_rows(recon_source_rows, vault_name)
    write_recon_sheet(workbook, recon_rows)

    usdt_header, usdt_rows = build_usdt_rows([extended_header] + filtered_rows)
    write_usdt_sheet(workbook, "USDT", usdt_header, usdt_rows, vault_name, trx_address, eth_address)
    apply_usdt_number_formats(workbook["USDT"], first_data_row=5)
    apply_usdt_colour_blocks(workbook["USDT"], header_row=4, first_data_row=5)

    trx_header, trx_rows = build_base_asset_rows([extended_header] + filtered_rows, TRX_SHEET_ASSETS)
    write_single_wallet_sheet(workbook, "TRX", trx_header, trx_rows, f"{vault_name} TRX Address", trx_address, "TRX")
    apply_base_asset_number_formats(workbook["TRX"], first_data_row=5)
    apply_base_asset_colour_blocks(workbook["TRX"], header_row=4, first_data_row=5)

    eth_header, eth_rows = build_base_asset_rows([extended_header] + filtered_rows, ETH_SHEET_ASSETS)
    write_single_wallet_sheet(workbook, "ETH", eth_header, eth_rows, f"{vault_name} ETH Address", eth_address, "ETH")
    apply_base_asset_number_formats(workbook["ETH"], first_data_row=5)
    apply_base_asset_colour_blocks(workbook["ETH"], header_row=4, first_data_row=5)

    write_standard_sheet(workbook, "Consolidated", extended_header, filtered_rows)

    # Ensure the output folder exists, then save the workbook.
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook.save(output_file)

    # Print a short summary so the user can see what the script did.
    print(f"Vault name: {vault_name}")
    print(f"TRX address: {trx_address}")
    print(f"ETH address: {eth_address}")
    print(f"Workbook created: {output_file}")
    print(f"Imported CSV files: {', '.join(str(path.name) for path in csv_files)}")
    print(f"Recon file used: {recon_file.name}")
    print(f"Completed rows kept: {len(completed_rows)}")
    print(f"Rows kept after asset filter: {len(filtered_rows)}")
    print(f"TRX rows kept: {len(trx_rows)}")
    print(f"ETH rows kept: {len(eth_rows)}")
    print(f"USDT rows kept: {len(usdt_rows)}")
    print(f"Recon rows kept: {len(recon_rows)}")


# This line means: only run the script automatically when this file is executed directly.
if __name__ == "__main__":
    main()
